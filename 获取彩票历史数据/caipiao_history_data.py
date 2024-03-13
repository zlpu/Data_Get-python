"""
功能说明：通过python获取彩票（大乐透）的历史开奖数据
数据源：https://datachart.500.com/dlt/history/history.shtml

注意：需要历史数据的时间区间确定最终的url,F12分析网页请求即可得出真实请求地址
# Pdata_url = "https://datachart.500.com/dlt/history/history.shtml"
# Pdata_url ="https://datachart.500.com/dlt/history/newinc/history.php?start=10000&end=24024"

# 七星彩  https://datachart.500.com/qxc/history/inc/history.php?limit=24025&start=00001&end=24026
"""

import requests,sys,io
from lxml import etree
import pandas as pd
from openpyxl import Workbook


sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

"""定义获取大乐透/双色球历史数据的函数（正常情况）"""


def get_history_data_D_S15(url):
    # 获取网页
    rs_1 = requests.get(url)  # 这一行使用了requests库中的get方法来向指定的url发送一个HTTP GET请求，然后将响应保存在rs_1变量中。这个响应对象包含了从指定网址获取的内容。
    html_1 = rs_1.text  # 这一行将响应对象rs_1的文本内容保存在html_1变量中。这个文本内容通常就是网页的HTML代码。

    # 使用lxml解析html
    html_tree = etree.HTML(html_1)  # 这一行使用了etree库中的HTML方法，将网页的HTML文本html_1转换为一个可供XPath解析的HTML树对象html_tree

    # 通过Xpath获取网页内容（根据前期分析网页，基本确定是table元素）
    html_tables = html_tree.xpath(
        '//*[@id="tdata"]')  # 这一行使用xpath方法从HTML树中提取数据。在这里，它使用XPath表达式'//*[@id="tdata"]'来获取id为"tdata"的元素，通常这种情况下会获取到一个或多个table元素，将结果保存在html_tables变量中。

    # 打印元素中table的内容，因为本次请求只涉及一个table元素，我们分析一个元素中的数据即可，这里可以不用遍历，直接使用html_tables[0]
    """遍历打印"""
    for html_table in html_tables:
        # 获取表格中所有的文本内容，用空格隔开
        text_content = ' '.join(html_table.itertext())
        # print(text_content)

        # 将数据转化为list,每15个空格为一行
        str_list = text_content.split()  # 将字符串按空格分割成列表
        # print(str_list)

        list1 = []  # 创建一个空列表，用于存放子列表
        sub_list = []  # 创建一个空列表，用于临时存放子列表的元素

        for item in str_list:
            if len(sub_list) < 15:  # 如果临时子列表的长度小于15,并且因为双色球以前有快乐星期八的数字，所以需要剔除数字0.
                sub_list.append(item)  # 将元素添加到临时子列表中
            else:  # 如果临时子列表的长度达到15
                list1.append(sub_list)  # 将临时子列表添加到结果列表中
                sub_list = [item]  # 重新创建一个临时子列表，将当前元素添加进去

        # 处理完最后一组元素
        if sub_list:
            list1.append(sub_list)

        """保存为文件"""
        # 创建一个新的工作簿
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        # 定义表头
        if int(type_chioce) == 1:
            ws.append(["【大乐透】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"])
            ws.append(["期号", "前区-1", "前区-2", "前区-3", "前区-4", "前区-5", "后区-1", "后区-2", "奖池奖金（元）",
                       "一等奖-注数", "一等奖-奖金", "二等奖-注数", "二等奖-奖金", "总投注金额", "开奖日期"])
            file_name = "【大乐透】"+str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"
        elif int(type_chioce) == 2:
            ws.append(["【双色球】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"])
            ws.append(["期号", "红球-1", "红球-2", "红球-3", "红球-4", "红球-5", "红球-6", "篮球-1",
                       "奖池奖金（元）",
                       "一等奖-注数", "一等奖-奖金", "二等奖-注数", "二等奖-奖金", "总投注金额", "开奖日期"])
            file_name = "【双色球】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"

        # 将数据写入工作表
        for row in list1:
            ws.append(row)
        wb.save(file_name + ".xlsx")

        # 终端打印
        # print(list1)
        print("总行数：" + str(len(list1)))

        # 遍历打印每一行
        for list_ in list1:
            print(list_)


"""定义获取双色球历史数据的函数（特殊情况：快乐星期八）"""


def get_history_data_D_S16(url, file_name_):
    # 获取网页
    rs_1 = requests.get(url)  # 这一行使用了requests库中的get方法来向指定的url发送一个HTTP GET请求，然后将响应保存在rs_1变量中。这个响应对象包含了从指定网址获取的内容。
    html_1 = rs_1.text  # 这一行将响应对象rs_1的文本内容保存在html_1变量中。这个文本内容通常就是网页的HTML代码。

    # 使用lxml解析html
    html_tree = etree.HTML(html_1)  # 这一行使用了etree库中的HTML方法，将网页的HTML文本html_1转换为一个可供XPath解析的HTML树对象html_tree

    # 通过Xpath获取网页内容（根据前期分析网页，基本确定是table元素）
    html_tables = html_tree.xpath(
        '//*[@id="tdata"]')  # 这一行使用xpath方法从HTML树中提取数据。在这里，它使用XPath表达式'//*[@id="tdata"]'来获取id为"tdata"的元素，通常这种情况下会获取到一个或多个table元素，将结果保存在html_tables变量中。

    # 打印元素中table的内容，因为本次请求只涉及一个table元素，我们分析一个元素中的数据即可，这里可以不用遍历，直接使用html_tables[0]
    """遍历打印"""
    for html_table in html_tables:
        # 获取表格中所有的文本内容，用空格隔开
        text_content = ' '.join(html_table.itertext())
        # print(text_content)

        # 将数据转化为list,每15个空格为一行
        str_list = text_content.split()  # 将字符串按空格分割成列表

        list1 = []  # 创建一个空列表，用于存放子列表
        sub_list = []  # 创建一个空列表，用于临时存放子列表的元素

        for item in str_list:
            if len(sub_list) < 16:  # 如果临时子列表的长度小于15,并且因为双色球以前有快乐星期八的数字，所以需要剔除数字0.
                sub_list.append(item)  # 将元素添加到临时子列表中
            else:  # 如果临时子列表的长度达到15
                list1.append(sub_list)  # 将临时子列表添加到结果列表中
                sub_list = [item]  # 重新创建一个临时子列表，将当前元素添加进去

        # 处理完最后一组元素
        if sub_list:
            list1.append(sub_list)

        """保存为文件"""
        # 创建一个新的工作簿
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        # 定义表头
        ws.append(["【双色球】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"])
        ws.append(
            ["期号", "红球-1", "红球-2", "红球-3", "红球-4", "红球-5", "红球-6", "篮球-1", "快乐星期八", "奖池奖金（元）",
             "一等奖-注数", "一等奖-奖金", "二等奖-注数", "二等奖-奖金", "总投注金额", "开奖日期"])

        # 将数据写入工作表
        for row in list1:
            ws.append(row)
        # 保存工作簿
        wb.save(file_name_ + ".xlsx")

        # 终端打印
        # print(list1)
        print("总行数：" + str(len(list1)))

        # 遍历打印每一行
        for list_ in list1:
            print(list_)


""" 定义获取七星彩\排列5历史数据的函数"""
def get_history_data_57(url):
    # 获取网页
    rs_1 = requests.get(url)  # 这一行使用了requests库中的get方法来向指定的url发送一个HTTP GET请求，然后将响应保存在rs_1变量中。这个响应对象包含了从指定网址获取的内容。
    rs_1.encoding = 'utf-8'
    html_1 = rs_1.text  # 这一行将响应对象rs_1的文本内容保存在html_1变量中。这个文本内容通常就是网页的HTML代码。

    # 使用lxml解析html
    html_tree = etree.HTML(html_1)  # 这一行使用了etree库中的HTML方法，将网页的HTML文本html_1转换为一个可供XPath解析的HTML树对象html_tree
    # 通过Xpath获取网页内容（根据前期分析网页，基本确定是table元素）
    html_tables = html_tree.xpath('//*[@id="container"]/div/div/div[2]')
    # 这一行使用xpath方法从HTML树中提取数据。在这里，它使用XPath表达式'//*[@id="tdata"]'来获取id为"tdata"的元素，通常这种情况下会获取到一个或多个table元素，将结果保存在html_tables变量中。

    # 打印元素中table的内容，因为本次请求只涉及一个table元素，我们分析一个元素中的数据即可，这里可以不用遍历，直接使用html_tables[0]
    """遍历打印"""
    for html_table in html_tables:
        # 获取表格中所有的文本内容，用空格隔开
        text_content = ' '.join(html_table.itertext())
        # print(text_content)

        # 将数据转化为list,每15个空格为一行
        str_list = text_content.split()[5:]  # 将字符串按空格分割成列表
        print(str_list)
        list1 = []  # 创建一个空列表，用于存放子列表
        sub_list = []  # 创建一个空列表，用于临时存放子列表的元素

        """每种彩票的数字个数不一样，换行的依据不一样"""

        if int(type_chioce) == 5:
            AtoB = 9
            file_name = "排列5" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"
        elif int(type_chioce) == 7:
            AtoB = 11
            file_name = "七星彩" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"

        for item in str_list:
            if len(sub_list) < AtoB:  # 如果临时子列表的长度小于10,并且因为双色球以前有快乐星期八的数字，所以需要剔除数字0.
                sub_list.append(item)  # 将元素添加到临时子列表中
            else:  # 如果临时子列表的长度达到15
                list1.append(sub_list)  # 将临时子列表添加到结果列表中
                sub_list = [item]  # 重新创建一个临时子列表，将当前元素添加进去

        # 处理完最后一组元素
        if sub_list:
            list1.append(sub_list)

        """保存为文件"""
        # 创建一个新的工作簿
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        # 定义表头

        if int(type_chioce) == 5:
            ws.append(["【排列5】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"])
            ws.append(
                ["期号", "中奖号-1", "中奖号-2", "中奖号-3", "中奖号-4", "中奖号-5", "总和",
                 "总销售额(元)", "开奖日期"])
            file_name = "【排列3】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"

        elif int(type_chioce) == 7:
            ws.append(["【七星彩】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"])
            ws.append(
                ["期号", "中奖号-1", "中奖号-2", "中奖号-3", "中奖号-4", "中奖号-5", "中奖号-6", "中奖号-7", "全加和",
                 "总销售额(元)", "开奖日期"])
            file_name = "【七星彩】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"

        # 将数据写入工作表
        for row in list1:
            ws.append(row)
        # 保存工作簿

        wb.save(file_name + ".xlsx")

        # 终端打印
        # print(list1)
        print("总行数：" + str(len(list1)))
        # 遍历打印每一行
        for list_ in list1:
            print(list_)




""" 定义获取排列3历史数据的函数"""
def get_history_data_3(url):
    # 获取网页
    rs_1 = requests.get(url)  # 这一行使用了requests库中的get方法来向指定的url发送一个HTTP GET请求，然后将响应保存在rs_1变量中。这个响应对象包含了从指定网址获取的内容。
    html_1 = rs_1.text  # 这一行将响应对象rs_1的文本内容保存在html_1变量中。这个文本内容通常就是网页的HTML代码。
    # 使用lxml解析html
    html_tree = etree.HTML(html_1)  # 这一行使用了etree库中的HTML方法，将网页的HTML文本html_1转换为一个可供XPath解析的HTML树对象html_tree
    # 通过Xpath获取网页内容（根据前期分析网页，基本确定是table元素）
    html_tables = html_tree.xpath('//*[@id="tablelist"]')
    # 这一行使用xpath方法从HTML树中提取数据。在这里，它使用XPath表达式'//*[@id="tdata"]'来获取id为"tdata"的元素，通常这种情况下会获取到一个或多个table元素，将结果保存在html_tables变量中。

    # 打印元素中table的内容，因为本次请求只涉及一个table元素，我们分析一个元素中的数据即可，这里可以不用遍历，直接使用html_tables[0]
    """遍历打印"""
    for html_table in html_tables:
        # 获取表格中所有的文本内容，用空格隔开
        text_content = ' '.join(html_table.itertext())
        # print(text_content)

        # 将数据转化为list,每15个空格为一行
        str_list = text_content.split()[14:]  # 将字符串按空格分割成列表
        print(str_list)
        list1 = []  # 创建一个空列表，用于存放子列表
        sub_list = []  # 创建一个空列表，用于临时存放子列表的元素

        """每种彩票的数字个数不一样，换行的依据不一样"""
        if int(type_chioce) == 3:
            AtoB = 13
        for item in str_list:
            if len(sub_list) < AtoB:  # 如果临时子列表的长度小于10,并且因为双色球以前有快乐星期八的数字，所以需要剔除数字0.
                sub_list.append(item)  # 将元素添加到临时子列表中
            else:  # 如果临时子列表的长度达到15
                list1.append(sub_list)  # 将临时子列表添加到结果列表中
                sub_list = [item]  # 重新创建一个临时子列表，将当前元素添加进去

        # 处理完最后一组元素
        if sub_list:
            list1.append(sub_list)

        """保存为文件"""
        # 创建一个新的工作簿
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        # 定义表头
        ws.append(["【排列3】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"])
        ws.append(
            ["期号","中奖号码-1","中奖号码-2","中奖号码-3","总和","总销售额(元)","直选-注数","直选-奖金","组选3-注数","直选-奖金","组选6-注数","直选-奖金","开奖日期"])

        # 将数据写入工作表
        for row in list1:
            ws.append(row)
        # 保存工作簿
        file_name = "【排列3】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"
        wb.save(file_name + ".xlsx")

        # 终端打印
        # print(list1)
        print("总行数：" + str(len(list1)))
        # 遍历打印每一行
        for list_ in list1:
            print(list_)



if __name__ == '__main__':
    while (1):
        print("请选择需要查询的彩票类型：1-大乐透；2-双色球；3-排列3;5-排列5;7-七星彩")
        type_chioce = input("请输入选项（1或2或7）：")
        print(
            "期数格式为5位数字，其中前两位为年份的后两位，后三位为当年的期数，比如24010表示2024年第10期")
        start_date = input("请输入历史数据开始期数：")
        end_date = input("请输入历史数据结束期数：")
        if (int(type_chioce) == 1):
            type_name = "dlt"
        elif (int(type_chioce) == 2):
            type_name = "ssq"
        elif (int(type_chioce) == 3):
            type_name = "pls"
        elif (int(type_chioce) == 5):
            type_name = "plw"
        elif (int(type_chioce) == 7):
            type_name = "qxc"
        else:
            print("请重新输入选项")

        # 结合每种彩票的特点，分支讨论，每种彩票使用一个函数去执行，双色球和大乐透因为数量大致相同，可以使用同一个函数。  """双色球考虑11070期之前含快乐星期八，需要分情况考虑"""

        if int(type_chioce) == 2:
            if int(start_date) <= 11070 and int(end_date) > 11070:
                base_url1 = "https://datachart.500.com/" + str(type_name) + "/history/newinc/history.php?start=" + str(
                    start_date) + "&end=11070"
                file_name_1 = "【双色球】" + str(start_date) + "期到11070期的历史开奖数据"
                get_history_data_D_S16(base_url1, file_name_1)
                file_name_2 = "11071期到" + str(end_date) + "期的历史开奖数据"
                base_url2 = "https://datachart.500.com/" + str(
                    type_name) + "/history/newinc/history.php?start=11071&end=" + str(end_date)
                get_history_data_D_S15(base_url2)
            elif int(start_date) > 11070:
                base_url = "https://datachart.500.com/" + str(type_name) + "/history/newinc/history.php?start=" + str(
                    start_date) + "&end=" + str(end_date)
                file_name = str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"
                get_history_data_D_S15(base_url)
            elif int(end_date) <= 11070:
                base_url = "https://datachart.500.com/" + str(type_name) + "/history/newinc/history.php?start=" + str(
                    start_date) + "&end=" + str(end_date)
                file_name = "【双色球】" + str(start_date) + "期到" + str(end_date) + "期的历史开奖数据"
                get_history_data_D_S16(base_url, file_name)

        # 七星彩\排列5历史数据
        elif int(type_chioce) == 7  or int(type_chioce) == 5:
            base_url = "https://datachart.500.com/" + str(type_name) + "/history/inc/history.php?limit="+str(int(int(end_date)-int(start_date)))+"&start=" + str(
                start_date) + "&end=" + str(end_date)
            get_history_data_57(base_url)

        elif int(type_chioce) ==3:
            base_url = "https://datachart.500.com/" + str(type_name) + "/history/inc/history.php?limit="+str(int(int(end_date)-int(start_date)))+"&start=" + str(
                start_date) + "&end=" + str(end_date)
            get_history_data_3(base_url)


        # 大乐透和双色球正常使用函数
        else:
            base_url = "https://datachart.500.com/" + str(type_name) + "/history/newinc/history.php?start=" + str(
                start_date) + "&end=" + str(end_date)
            get_history_data_D_S15(base_url)
